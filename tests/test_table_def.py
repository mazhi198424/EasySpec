from openpyxl import Workbook
from src.models import TableInfo, ColumnInfo, ForeignKeyInfo
from src.generators.table_def import generate_table_def
from src.generators.styles import create_styles


def test_generate_table_def():
    wb = Workbook()
    ws = wb.active
    tables = [
        TableInfo(
            name="employees", logical_name="社員情報",
            columns=[
                ColumnInfo(name="id", logical_name="ID", sql_type="BIGINT", type_name="BIGINT",
                          nullable=False, key_type="PK", description="主キー"),
                ColumnInfo(name="employee_no", logical_name="社員番号", sql_type="VARCHAR(20)", type_name="VARCHAR",
                          length=20, nullable=False, key_type="UNIQUE"),
                ColumnInfo(name="department_id", logical_name="部署ID", sql_type="BIGINT", type_name="BIGINT"),
            ],
            foreign_keys=[ForeignKeyInfo(column="department_id", ref_table="departments", ref_column="id")],
        ),
    ]
    styles = create_styles()
    generate_table_def(ws, tables, styles)
    assert "社員情報" in str(ws['A1'].value)
    assert ws.cell(row=2, column=1).value == "No"
    assert ws.cell(row=3, column=1).value == 1
