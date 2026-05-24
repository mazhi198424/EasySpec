from openpyxl import Workbook
from src.models import SystemInfo
from src.generators.cover import generate_cover
from src.generators.styles import create_styles


def test_generate_cover_creates_sheet():
    wb = Workbook()
    ws = wb.active
    sys_info = SystemInfo(name="TestApp", version="1.0.0", packaging="war")
    styles = create_styles()
    generate_cover(ws, sys_info, styles)

    assert ws['A1'].value == "システム詳細設計書"
    assert ws['B3'].value == "TestApp"
    assert ws['B4'].value == "1.0.0"
    assert ws['A7'].value == "版数"
    assert ws['B7'].value == "更新日"
    assert ws['C7'].value == "更新内容"
    assert ws['D7'].value == "作成者"
    assert ws['A8'].value == "1.0"
