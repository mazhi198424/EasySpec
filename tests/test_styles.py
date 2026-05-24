from openpyxl import Workbook
from src.generators.styles import (
    create_styles, apply_header_style, apply_body_style,
    apply_title_style
)


def test_create_styles_returns_dict():
    styles = create_styles()
    assert 'header_font' in styles
    assert 'body_font' in styles
    assert 'title_font' in styles
    assert 'header_fill' in styles
    assert 'thin_border' in styles
    assert 'header_alignment' in styles
    assert 'body_alignment' in styles


def test_apply_header_style():
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'Test'
    styles = create_styles()
    apply_header_style(ws['A1'], styles)
    assert ws['A1'].font.bold is True


def test_apply_body_style():
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'Test'
    styles = create_styles()
    apply_body_style(ws['A1'], styles)
    assert ws['A1'].font.size == 10


def test_apply_title_style():
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'Title'
    styles = create_styles()
    apply_title_style(ws['A1'], styles)
    assert ws['A1'].font.size == 14
    assert ws['A1'].font.bold is True
