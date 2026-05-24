from openpyxl import Workbook
from src.generators.business_flow import generate_business_flow
from src.generators.styles import create_styles


def test_generate_business_flow():
    wb = Workbook()
    ws = wb.active
    styles = create_styles()
    generate_business_flow(ws, styles)
    assert ws['A1'].value == "業務フロー"
    assert "経費精算フロー" in str(ws['A3'].value)
