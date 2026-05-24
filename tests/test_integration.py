# tests/test_integration.py
import os
import tempfile
from generate_shiyosho import generate


def test_generate_creates_xlsx(visiondemo_path):
    with tempfile.TemporaryDirectory() as tmpdir:
        output_path = os.path.join(tmpdir, "test_output.xlsx")
        generate(visiondemo_path, output_path)

        assert os.path.exists(output_path)

        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        sheet_names = wb.sheetnames

        assert "表紙" in sheet_names
        assert "画面一覧" in sheet_names
        assert "テーブル定義" in sheet_names
        assert "ER図" in sheet_names
        assert "API仕様書" in sheet_names
        assert "業務フロー" in sheet_names

        # Check screen spec sheets exist
        screen_sheets = [s for s in sheet_names if s.startswith("SCR-")]
        assert len(screen_sheets) >= 8

        wb.close()
