from openpyxl import Workbook
from src.models import ScreenInfo
from src.generators.screen_list import generate_screen_list
from src.generators.styles import create_styles


def test_generate_screen_list():
    wb = Workbook()
    ws = wb.active
    screens = [
        ScreenInfo(id="SCR-001", name="Main", url="/", jsp_file="index.jsp", controller="IndexController"),
        ScreenInfo(id="SCR-002", name="Org", url="/org/page", jsp_file="fragments/org-tree.jsp", controller="OrgController"),
    ]
    styles = create_styles()
    generate_screen_list(ws, screens, styles)
    assert ws['A1'].value == "画面一覧"
    headers = [ws.cell(row=2, column=c).value for c in range(1, 6)]
    assert headers == ["画面ID", "画面名", "URL", "JSPファイル", "機能概要"]
    assert ws.cell(row=3, column=1).value == "SCR-001"
