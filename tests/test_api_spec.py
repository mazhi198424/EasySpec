from openpyxl import Workbook
from src.models import ApiEndpoint, ApiParam
from src.generators.api_spec import generate_api_spec
from src.generators.styles import create_styles


def test_generate_api_spec():
    wb = Workbook()
    ws = wb.active
    endpoints = [
        ApiEndpoint(method="GET", url="/employee/api/search", controller="EmployeeController",
                    request_params=[ApiParam(name="name", location="query", required=False)],
                    response_type="Map<String, Object>", description="社員検索"),
    ]
    styles = create_styles()
    generate_api_spec(ws, endpoints, styles)
    assert ws['A1'].value == "API仕様書"
    assert ws.cell(row=2, column=1).value == "No"
    assert ws.cell(row=3, column=3).value == "GET"
