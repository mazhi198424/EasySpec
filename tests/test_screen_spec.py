from openpyxl import Workbook
from src.models import ScreenInfo, ScreenField, ScreenButton, TableColumnDef
from src.generators.screen_spec import generate_screen_specs
from src.generators.styles import create_styles


def test_generate_screen_specs():
    wb = Workbook()
    screens = [
        ScreenInfo(
            id="SCR-003", name="社員管理", url="/employee/page",
            jsp_file="fragments/employee-list.jsp", controller="EmployeeController",
            layout_description="検索エリア + 一覧テーブル",
            fields=[ScreenField(name="emp-search-name", field_type="text")],
            buttons=[ScreenButton(name="emp-search-btn", action="検索")],
            table_columns=[TableColumnDef(key="employeeNo", label="社員番号", width="90px")],
        ),
    ]
    styles = create_styles()
    generate_screen_specs(wb, screens, styles)
    ws = wb["SCR-003_社員管理"]
    assert ws is not None
    assert ws['A1'].value == "画面仕様書"
    assert "SCR-003" in str(ws['B2'].value)
