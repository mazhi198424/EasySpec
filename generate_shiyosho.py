# generate_shiyosho.py
import os
import sys
from openpyxl import Workbook
from src.project_reader import ProjectReader
from src.generators.styles import create_styles
from src.generators.cover import generate_cover
from src.generators.screen_list import generate_screen_list
from src.generators.table_def import generate_table_def
from src.generators.er_diagram import generate_er_diagram, build_relationships
from src.generators.screen_spec import generate_screen_specs
from src.generators.api_spec import generate_api_spec
from src.generators.business_flow import generate_business_flow


def generate(project_path: str, output_path: str):
    reader = ProjectReader(project_path)
    styles = create_styles()

    wb = Workbook()

    # Sheet 1: 表紙
    ws1 = wb.active
    ws1.title = "表紙"
    sys_info = reader.read_system_info()
    generate_cover(ws1, sys_info, styles)

    # Collect data
    tables = reader.read_tables()
    screens = reader.read_screens()
    endpoints, _ = reader.read_api_endpoints()
    relationships = build_relationships(tables)

    # Sheet 2: 画面一覧
    ws2 = wb.create_sheet("画面一覧")
    generate_screen_list(ws2, screens, styles)

    # Sheet 3: テーブル定義
    ws3 = wb.create_sheet("テーブル定義")
    generate_table_def(ws3, tables, styles)

    # Sheet 4: ER図
    ws4 = wb.create_sheet("ER図")
    generate_er_diagram(ws4, tables, relationships, styles)

    # Sheets 5-13: 画面仕様 (one per screen)
    generate_screen_specs(wb, screens, styles)

    # Sheet 14: API仕様書
    ws_api = wb.create_sheet("API仕様書")
    generate_api_spec(ws_api, endpoints, styles)

    # Sheet 15: 業務フロー
    ws_flow = wb.create_sheet("業務フロー")
    generate_business_flow(ws_flow, styles)

    # Save
    os.makedirs(os.path.dirname(output_path) or '.', exist_ok=True)
    wb.save(output_path)
    print(f"式様書を生成しました: {output_path}")


def main():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    default_project = os.path.join(script_dir, '..', 'VisionDemo')
    default_output = os.path.join(default_project, 'docs', 'shiyosho',
                                  'VisionDemo_式様書_v1.0.xlsx')

    project_path = sys.argv[1] if len(sys.argv) > 1 else default_project
    output_path = sys.argv[2] if len(sys.argv) > 2 else default_output

    project_path = os.path.abspath(project_path)
    output_path = os.path.abspath(output_path)

    if not os.path.isdir(project_path):
        print(f"エラー: プロジェクトパスが見つかりません: {project_path}")
        sys.exit(1)

    generate(project_path, output_path)


if __name__ == '__main__':
    main()
