from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def create_styles() -> dict:
    return {
        'title_font': Font(name='Yu Gothic', size=14, bold=True),
        'header_font': Font(name='Yu Gothic', size=10, bold=True),
        'body_font': Font(name='Yu Gothic', size=10),
        'small_font': Font(name='Yu Gothic', size=9),
        'header_fill': PatternFill(start_color='DCE6F1', end_color='DCE6F1', fill_type='solid'),
        'thin_border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        ),
        'header_alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'body_alignment': Alignment(vertical='center', wrap_text=True),
        'center_alignment': Alignment(horizontal='center', vertical='center'),
    }


def apply_header_style(cell, styles):
    cell.font = styles['header_font']
    cell.fill = styles['header_fill']
    cell.border = styles['thin_border']
    cell.alignment = styles['header_alignment']


def apply_body_style(cell, styles):
    cell.font = styles['body_font']
    cell.border = styles['thin_border']
    cell.alignment = styles['body_alignment']


def apply_title_style(cell, styles):
    cell.font = styles['title_font']
    cell.alignment = Alignment(horizontal='left', vertical='center')


def write_header_row(ws, row, headers, styles, start_col=1):
    for i, header in enumerate(headers):
        cell = ws.cell(row=row, column=start_col + i, value=header)
        apply_header_style(cell, styles)


def write_data_row(ws, row, data, styles, start_col=1):
    for i, value in enumerate(data):
        cell = ws.cell(row=row, column=start_col + i, value=value)
        apply_body_style(cell, styles)
